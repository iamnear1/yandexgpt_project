from openpyxl import Workbook

from lib.parser import *
from lib.reviewers import FullTaskReviewer

if __name__ == "__main__":

    wb = Workbook()
    ws = wb.active

    results = [[]]

    all_works = get_notebooks_filenames_from_directory("works")

    # не сделанная работа
    orig_work_path = "../data/test/Домашнее задание 4 (1).ipynb"

    parsed_notebooks, marks = [], []
    for path in all_works:
        p, m = parsing_pipeline(path, orig_work_path, MergeKind.BY_CHANGE, 3)
        parsed_notebooks.append(p)
        marks.append(m)

    TASKS = 1

    ws.append(["название файла", "оценка оценки yagpt"])

    reviewer: FullTaskReviewer = FullTaskReviewer(...)

    for i in range(1, len(all_works)):
        filepath = all_works[i]
        for (j, task) in enumerate(parsed_notebooks[i][0:TASKS]):
            ws.cell(row=i + 1, column=1).value = filepath

            answer = reviewer.review(task)
            ws.cell(row=i + 1, column=3 + j).value = answer
